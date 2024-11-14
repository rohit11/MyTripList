import requests
import json
from deepdiff import DeepDiff
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from collections import defaultdict

# Load JSON data from a URL and remove 'id' fields
def load_json_from_url(url):
    response = requests.get(url)
    response.raise_for_status()  # Check for request errors
    data = response.json()

    def remove_ids(obj):
        if isinstance(obj, dict):
            return {k: remove_ids(v) for k, v in obj.items() if k != "id"}
        elif isinstance(obj, list):
            return [remove_ids(item) for item in obj]
        else:
            return obj

    return remove_ids(data)

# Categorize changes by parent key
def categorize_changes_by_parent_key(old_data, new_data):
    diff = DeepDiff(old_data, new_data, ignore_order=True)
    changes_by_parent_key = defaultdict(list)

    def get_parent_key(path):
        return path.split('.')[0] if '.' in path else path

    if 'dictionary_item_added' in diff:
        for item in diff['dictionary_item_added']:
            parent_key = get_parent_key(item)
            changes_by_parent_key[parent_key].append({
                'Key': item, 'Type': 'New', 'Value': diff['dictionary_item_added'][item]
            })

    if 'values_changed' in diff:
        for item, change in diff['values_changed'].items():
            parent_key = get_parent_key(item)
            changes_by_parent_key[parent_key].append({
                'Key': item, 
                'Type': 'Updated', 
                'Old Value': change['old_value'], 
                'New Value': change['new_value']
            })

    if 'dictionary_item_removed' in diff:
        for item in diff['dictionary_item_removed']:
            parent_key = get_parent_key(item)
            changes_by_parent_key[parent_key].append({
                'Key': item, 'Type': 'Deleted', 'Value': diff['dictionary_item_removed'][item]
            })

    return changes_by_parent_key

# Export changes to Excel with color coding
def export_changes_to_excel(changes_by_parent_key, output_file):
    wb = Workbook()
    del wb['Sheet']  # Remove default sheet

    # Define fill colors for types of changes
    new_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for New
    updated_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for Updated
    deleted_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for Deleted

    for parent_key, changes in changes_by_parent_key.items():
        # Convert the changes list to a DataFrame
        df = pd.DataFrame(changes)
        
        # Create a new sheet for each parent key
        ws = wb.create_sheet(title=parent_key)

        # Write DataFrame to sheet
        for row_idx, row in enumerate(df.to_records(index=False), start=2):  # start from row 2
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Apply headers
        for col_idx, header in enumerate(df.columns, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Apply color coding based on 'Type'
        for row_idx, change_type in enumerate(df['Type'], start=2):
            if change_type == 'New':
                for cell in ws[row_idx:row_idx+1]:  # Full row
                    cell.fill = new_fill
            elif change_type == 'Updated':
                for cell in ws[row_idx:row_idx+1]:
                    cell.fill = updated_fill
            elif change_type == 'Deleted':
                for cell in ws[row_idx:row_idx+1]:
                    cell.fill = deleted_fill

    # Save workbook
    wb.save(output_file)
    print(f"Changes exported to {output_file}")

# Example usage
old_url = 'https://example.com/path_to_old_json.json'  # Replace with actual URL
new_url = 'https://example.com/path_to_new_json.json'  # Replace with actual URL

old_data = load_json_from_url(old_url)
new_data = load_json_from_url(new_url)

# Categorize changes
changes = categorize_changes_by_parent_key(old_data, new_data)

# Export to Excel
export_changes_to_excel(changes, 'json_differences.xlsx')